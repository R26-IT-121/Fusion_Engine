"""
Batch analysis of an uploaded transaction file.

This is how a bank actually uses the system: a day's transactions arrive as a
file, not one at a time through a form.

Parsing accepts CSV and Excel in the PaySim schema. Column matching is
case-insensitive and tolerant of the separator style a spreadsheet export
produces (nameOrig / name_orig / NAMEORIG), because a file that fails to load
for a punctuation reason is the most common way a demo stalls.

An isFraud column, when present, is read as ground truth and never as input to
scoring. It is used only to score detection afterwards.

Per-transaction LLM narration is deliberately not performed. Report generation
takes seconds per transaction, so a 300-row file would take the better part of
an hour. Scoring, fusion and typology retrieval run for every row; a narrative
is generated afterwards for the highest-risk rows only.
"""

import csv
import io
import logging
import math
from dataclasses import dataclass, field
from typing import Any, Iterator, Optional

logger = logging.getLogger(__name__)

MAX_ROWS = 5_000
MAX_BYTES = 10 * 1024 * 1024  # 10 MB

REQUIRED = ("step", "type", "amount", "nameorig", "namedest")
OPTIONAL_NUMERIC = (
    "oldbalanceorg",
    "newbalanceorig",
    "oldbalancedest",
    "newbalancedest",
)

VALID_TYPES = {"TRANSFER", "CASH_OUT", "CASH_IN", "PAYMENT", "DEBIT"}


class BatchError(Exception):
    """Raised for a file the user needs to fix, with a message that says how."""


@dataclass
class ParsedRow:
    index: int          # 1-based row number in the source file, for error reports
    transaction: dict
    is_fraud_label: Optional[int] = None   # ground truth, if the file carried it
    typology_label: Optional[str] = None


@dataclass
class BatchSummary:
    total: int = 0
    analysed: int = 0
    skipped: int = 0          # malformed rows, rejected before scoring
    unscored: int = 0         # no model was reachable, so no verdict was reached
    by_classification: dict = field(default_factory=dict)
    alerts: int = 0

    # Populated only when the file carried ground-truth labels
    has_labels: bool = False
    true_positive: int = 0
    false_positive: int = 0
    true_negative: int = 0
    false_negative: int = 0

    def metrics(self) -> dict:
        if not self.has_labels:
            return {}
        tp, fp, tn, fn = (
            self.true_positive,
            self.false_positive,
            self.true_negative,
            self.false_negative,
        )
        precision = tp / (tp + fp) if (tp + fp) else None
        recall = tp / (tp + fn) if (tp + fn) else None
        f1 = (
            2 * precision * recall / (precision + recall)
            if precision and recall
            else None
        )
        accuracy = (tp + tn) / (tp + fp + tn + fn) if (tp + fp + tn + fn) else None
        return {
            "true_positive": tp,
            "false_positive": fp,
            "true_negative": tn,
            "false_negative": fn,
            "precision": precision,
            "recall": recall,
            "f1": f1,
            "accuracy": accuracy,
        }


def _normalise(name: str) -> str:
    """Fold a header to a comparable form: 'name_Orig ' -> 'nameorig'."""
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def _to_float(value: Any, column: str, row_no: int) -> float:
    if value is None or value == "":
        return 0.0
    try:
        # Spreadsheets export thousands separators and currency symbols.
        cleaned = str(value).replace(",", "").replace("$", "").strip()
        f = float(cleaned)
    except (TypeError, ValueError):
        raise BatchError(f"Row {row_no}: '{column}' is not a number (got {value!r}).")
    if math.isnan(f) or math.isinf(f):
        raise BatchError(f"Row {row_no}: '{column}' is not a finite number.")
    return max(0.0, f)


def _to_int(value: Any, column: str, row_no: int) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        raise BatchError(f"Row {row_no}: '{column}' is not a whole number (got {value!r}).")


def _read_csv(data: bytes) -> tuple[list[str], Iterator[list]]:
    # utf-8-sig strips the BOM Excel writes, which would otherwise corrupt the
    # first header and make 'step' unfindable.
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = data.decode("latin-1")
        except UnicodeDecodeError:
            raise BatchError("The file is not readable as text. Save it as UTF-8 CSV.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # fall back to comma

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header = next(reader)
    except StopIteration:
        raise BatchError("The file is empty.")
    return header, reader


def _read_excel(data: bytes) -> tuple[list[str], Iterator[list]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise BatchError(
            "Excel support is unavailable on the server. Save the file as CSV instead."
        )

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise BatchError("That file could not be opened as a spreadsheet.")

    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [c for c in next(rows)]
    except StopIteration:
        raise BatchError("The spreadsheet's first sheet is empty.")
    return [str(h) if h is not None else "" for h in header], rows


def parse_file(filename: str, data: bytes) -> list[ParsedRow]:
    """Parse an uploaded file into transactions. Raises BatchError with guidance."""
    if not data:
        raise BatchError("The file is empty.")
    if len(data) > MAX_BYTES:
        raise BatchError(
            f"The file is {len(data) / 1024 / 1024:.1f} MB. The limit is "
            f"{MAX_BYTES // 1024 // 1024} MB — split it or trim unused columns."
        )

    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        header, rows = _read_excel(data)
    elif lower.endswith((".csv", ".txt", ".tsv")):
        header, rows = _read_csv(data)
    elif lower.endswith(".xls"):
        raise BatchError(
            "The legacy .xls format is not supported. Save it as .xlsx or .csv."
        )
    else:
        raise BatchError(
            "Upload a .csv or .xlsx file. Other formats cannot be read."
        )

    index = {_normalise(h): i for i, h in enumerate(header) if str(h).strip()}

    missing = [c for c in REQUIRED if c not in index]
    if missing:
        raise BatchError(
            "The file is missing required column(s): "
            + ", ".join(missing)
            + ". Expected the PaySim schema: step, type, amount, nameOrig, "
            "nameDest, oldbalanceOrg, newbalanceOrig, oldbalanceDest, newbalanceDest. "
            f"Found: {', '.join(str(h) for h in header if str(h).strip()) or 'nothing'}."
        )

    def cell(row: list, key: str):
        i = index.get(key)
        if i is None or i >= len(row):
            return None
        return row[i]

    parsed: list[ParsedRow] = []

    for offset, row in enumerate(rows, start=2):  # row 1 is the header
        if row is None:
            continue
        values = list(row)
        if not any(str(v).strip() for v in values if v is not None):
            continue  # blank line

        if len(parsed) >= MAX_ROWS:
            raise BatchError(
                f"The file has more than {MAX_ROWS:,} rows. Split it into smaller batches."
            )

        tx_type = str(cell(values, "type") or "").strip().upper()
        if tx_type not in VALID_TYPES:
            raise BatchError(
                f"Row {offset}: '{tx_type or '(blank)'}' is not a transaction type. "
                f"Expected one of {', '.join(sorted(VALID_TYPES))}."
            )

        step = _to_int(cell(values, "step"), "step", offset)
        # The model was trained on a 744-hour simulation month; clamp rather than
        # reject, so a file using absolute hours still runs.
        step = min(max(step, 1), 744)

        transaction = {
            "step": step,
            "type": tx_type,
            "amount": _to_float(cell(values, "amount"), "amount", offset),
            "nameOrig": str(cell(values, "nameorig") or "").strip(),
            "nameDest": str(cell(values, "namedest") or "").strip(),
            "oldbalanceOrg": _to_float(cell(values, "oldbalanceorg"), "oldbalanceOrg", offset),
            "newbalanceOrig": _to_float(cell(values, "newbalanceorig"), "newbalanceOrig", offset),
            "oldbalanceDest": _to_float(cell(values, "oldbalancedest"), "oldbalanceDest", offset),
            "newbalanceDest": _to_float(cell(values, "newbalancedest"), "newbalanceDest", offset),
            "isFlaggedFraud": _to_int(cell(values, "isflaggedfraud"), "isFlaggedFraud", offset),
        }

        if not transaction["nameOrig"] or not transaction["nameDest"]:
            raise BatchError(f"Row {offset}: nameOrig and nameDest cannot be blank.")

        label = cell(values, "isfraud")
        parsed.append(
            ParsedRow(
                index=offset,
                transaction=transaction,
                is_fraud_label=_to_int(label, "isFraud", offset) if label not in (None, "") else None,
                typology_label=(str(cell(values, "typology")).strip() or None)
                if cell(values, "typology") not in (None, "")
                else None,
            )
        )

    if not parsed:
        raise BatchError("No data rows were found beneath the header.")

    return parsed


class UpstreamCircuit:
    """
    Stops re-dialling a model API that has proved unreachable.

    Without this, a file is catastrophically slow when a model is down: each row
    waits out the full per-call timeout, so 330 rows against three unreachable
    models takes roughly half an hour rather than seconds. The first batch test
    hit exactly that and had to be killed.

    After `threshold` consecutive failures a modality is treated as down for the
    remainder of the batch and skipped outright. Its score is imputed and the
    fused confidence penalised, which is the same path a single-transaction
    request takes — so results stay consistent, they just arrive quickly.

    Scoped to one batch: a new upload dials again from scratch, because a model
    that was down a minute ago may be back.
    """

    def __init__(self, threshold: int = 3):
        self.threshold = threshold
        self._consecutive_failures: dict[str, int] = {}
        self._open: set[str] = set()

    def is_open(self, modality: str) -> bool:
        """True when this modality should be skipped for the rest of the batch."""
        return modality in self._open

    def record(self, modality: str, available: bool) -> None:
        if available:
            self._consecutive_failures[modality] = 0
            self._open.discard(modality)
            return

        count = self._consecutive_failures.get(modality, 0) + 1
        self._consecutive_failures[modality] = count
        if count >= self.threshold and modality not in self._open:
            self._open.add(modality)
            logger.warning(
                f"Upstream '{modality}' failed {count} times consecutively — "
                f"skipping it for the remainder of this batch."
            )

    @property
    def skipped(self) -> list[str]:
        return sorted(self._open)


def update_summary(
    summary: BatchSummary,
    classification: str,
    alerted: bool,
    label: Optional[int],
) -> None:
    summary.analysed += 1
    summary.by_classification[classification] = (
        summary.by_classification.get(classification, 0) + 1
    )
    if alerted:
        summary.alerts += 1

    if label is None:
        return

    summary.has_labels = True
    if label == 1 and alerted:
        summary.true_positive += 1
    elif label == 1 and not alerted:
        summary.false_negative += 1
    elif label == 0 and alerted:
        summary.false_positive += 1
    else:
        summary.true_negative += 1
